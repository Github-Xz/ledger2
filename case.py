import re
import time

from openpyxl import load_workbook

filename = '执法，采集，公示案件文书统计.xlsx'
wb = load_workbook(filename)
ws = wb.active


def get_case(site):
    """读取指定的sheet表"""
    sheet_ranges = wb['1案件总数统计']
    max_row = sheet_ranges.max_row

    row = 1
    hz_cases = []  # 惠州案件总数统计
    ns_cases = []  # 南沙案件总是统计

    sft_cases = []  # 司法厅案件总数统计
    while row <= max_row:
        if sheet_ranges['A' + str(row)].value == site:  # 以后修改 先对列名称进行判断 然后从最大列进行遍历 判断非空后取案件总数的列表
            row += 1

            hz_cases.append(sheet_ranges['J' + str(row)].value)  # 检查案件
            hz_cases.append(sheet_ranges['I' + str(row)].value)
            hz_cases.append(sheet_ranges['K' + str(row)].value)
            hz_cases.append(sheet_ranges['L' + str(row)].value)
            hz_cases.append(sheet_ranges['N' + str(row)].value)
            hz_cases.append(sheet_ranges['M' + str(row)].value)  # 检查文书
            hz_cases.append(sheet_ranges['O' + str(row)].value)
            hz_cases.append(sheet_ranges['P' + str(row)].value)
            break
            # print('惠州案件总数成功读取')
            #print(hz_cases)

        elif sheet_ranges['A' + str(row)].value == site:
            row += 1

            ns_cases.append(sheet_ranges['K' + str(row)].value)  # 检查案件
            ns_cases.append(sheet_ranges['J' + str(row)].value)
            ns_cases.append(sheet_ranges['L' + str(row)].value)  # 检查文书
            ns_cases.append(sheet_ranges['M' + str(row)].value)
            ns_cases.append(sheet_ranges['O' + str(row)].value)
            ns_cases.append(sheet_ranges['N' + str(row)].value)  # 处罚文书
            ns_cases.append(sheet_ranges['P' + str(row)].value)
            ns_cases.append(sheet_ranges['Q' + str(row)].value)
            break
           # print(ns_cases)

        elif sheet_ranges['A' + str(row)].value == site:
            row += 1

            sft_cases.append(sheet_ranges['J' + str(row)].value)
            sft_cases.append(sheet_ranges['I' + str(row)].value)
            sft_cases.append(sheet_ranges['K' + str(row)].value)
            sft_cases.append(sheet_ranges['L' + str(row)].value)
            sft_cases.append(sheet_ranges['N' + str(row)].value)
            sft_cases.append(sheet_ranges['M' + str(row)].value)
            sft_cases.append(sheet_ranges['O' + str(row)].value)
            sft_cases.append(sheet_ranges['P' + str(row)].value)
            # print('司法厅案件总数成功读取')
            #print(sft_cases)
            break
        else:
            row += 1

    if site == '南沙':
        print('南沙案件总数统计')
        print(ns_cases)
        return ns_cases
    elif site == '惠州':
        print('惠州案件总数统计')
        print(hz_cases)
        return hz_cases
    elif site == '司法厅':
        print('司法厅案件总数统计')
        print(sft_cases)
        return sft_cases


""""访问量处理"""


def get_page_view(site):  # 访问量处理     site为试点单位
    """读取指定的sheet表"""
    sheet_ranges = wb['04、访问量统计']
    print('读取')
    row = 1

    while True:

        if sheet_ranges['A' + str(row)].value == '执法平台访问量统计':
            row += 1
            dv = sheet_ranges['J' + str(row)].value  # 当天访问量
            uv = sheet_ranges['J' + str(row + 1)].value  # 当天访问数

            ns = []
            hz = []
            ns_dv = int(dv * 3 / 4)
            hz_dv = int(dv - ns_dv)
            ns_uv = int(uv * 3 / 4)
            hz_uv = int(uv - ns_uv)

            ns.append(ns_uv)  # 登录账号 登录次数
            ns.append(ns_dv)

            hz.append(hz_uv)
            hz.append(hz_dv)

            # hz.append(int(hz_uv * 0.1))
            # hz.append(int(hz_uv * 0.2))
            # hz.append(int(hz_uv * 0.2))
            # hz.append(int(hz_uv * 0.25))
            # hz.append(int(hz_uv * 0.25))
            #
            # hz.append(int(hz_dv * 0.1))
            # hz.append(int(hz_dv * 0.2))
            # hz.append(int(hz_dv * 0.2))
            # hz.append(int(hz_dv * 0.25))
            # hz.append(int(hz_dv * 0.25))
            break
        row += 1

    if site == "南沙":
        print(ns)
        return ns
    if site == "惠州":
        print(hz)
        return hz


def matching_case(site):  # 检查今天有没有数据，返回当日案件总数
    sheet_ranges = wb['1案件总数统计']
    max_row = sheet_ranges.max_row
    t = time.strftime("%Y%m%d", time.localtime())
    a = 1
    i = 2
    col = 1
    matching = []

    while a <= max_row:
        if sheet_ranges.cell(column=col, row=a).value == site:
            var = str(sheet_ranges.cell(column=col, row=a + 1).value)
            var = str(var[0:10])
            p = re.findall(r"\d+", var)
            p1 = p[0] + p[1] + p[2]
            if p1 == t:
                while i <= 7:
                    matching.append(sheet_ranges.cell(column=i, row=a + 1).value)
                    i += 1

        a += 1
    try:
        u = matching.pop(1)
        o = matching.pop(3)
        matching.insert(0, u)
        matching.insert(3, o)

    except IndexError:
        print(site + '今日没有数据')

    print('时间' + p1)
    return matching
