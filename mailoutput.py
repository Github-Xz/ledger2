from openpyxl import load_workbook

filename = '南沙区综合行政执法局执法平台试点台账.xlsx'
wb = load_workbook(filename)
sheet_ranges = wb['统计汇总表']


def loop():
    pass


def mailoutput(sheet_ranges, site, user_input):
    max_row = sheet_ranges.max_row
    max_col = sheet_ranges.max_column
    a = int(user_input[0]) + int(user_input[1]) + int(user_input[2])  # 今日新增案件
    b = int(user_input[3]) + int(user_input[4]) + int(user_input[5])  # 今日新增文书
    active = True

    while max_row != 0 and active == True:
        i = 1
        while i <= max_col:
            if sheet_ranges.cell(column=i, row=max_row).value == '有效数':
                c = sheet_ranges.cell(column=i + 4, row=max_row - 2).value  # 累计办理案件
                d = sheet_ranges.cell(column=i + 8, row=max_row - 2).value  # 累计办理文书
                e = sheet_ranges.cell(column=i + 8, row=max_row - 1).value  # 累计删除文书
                f = sheet_ranges.cell(column=i + 8, row=max_row).value  # 现有文书
                g = sheet_ranges.cell(column=i + 4, row=max_row - 1).value  # 删除案件总数
                h = sheet_ranges.cell(column=i + 4, row=max_row).value  # 现有案件

                # print(
                #     '各位领导晚上好！执法平台' + str(site) +
                #     '今日使用情况如下：今日新增案件' + str(a) +
                #     '宗，累计办理案件' + str(sheet_ranges.cell(column=i + 4, row=max_row - 2).value) +
                #     '宗；今日新增文书' + str(b) +
                #     '份，累计办理文书' + str(sheet_ranges.cell(column=i + 8,
                #                                        row=max_row - 2).value) + '份；今日移动端新增案件0宗，新增文书0份；今日删除案件0宗，今日删除文书0'
                #                                                                  '份；累计删除文书总数' + str(
                #         sheet_ranges.cell(column=i + 8, row=max_row - 1).value) + '份，'
                #                                                                   '现有文书' + str(
                #         sheet_ranges.cell(column=i + 8, row=max_row).value) + '份（其中移动端62份）；'
                #                                                               '累计删除案件总数' + str(
                #         sheet_ranges.cell(column=i + 4, row=max_row - 1).value) + '宗，现有案件' + str(
                #         sheet_ranges.cell(column=i + 4, row=max_row).value) + '宗（其中移动端42'
                #                                                               '宗）。目前系统操作问题都已解决，系统需求问题已经提交给产品记录。问题记录台账及运营数量统计及文书统计已经记录完成，现已发送附件给各位领导，请查阅。 '
                # )
                active = False
                break
            else:
                i += 1
        max_row -= 1
    # print(
    #     '各位领导晚上好！执法平台' + str(site) +
    #     '今日使用情况如下：今日新增案件' + str(a) +
    #     '宗，累计办理案件' + str(c) +
    #     '宗；今日新增文书' + str(b) +
    #     '份，累计办理文书' + str(d) + '份；今日移动端新增案件0宗，新增文书0份；今日删除案件0宗，今日删除文书0'
    #                           '份；累计删除文书总数' + str(e) + '份，' '现有文书' + str(f) + '份（其中移动端62份）；'
    #                                                                          '累计删除案件总数' + str(g) + '宗，现有案件' + str(
    #         h) + '宗（其中移动端42''宗）。'
    #              '目前系统操作问题都已解决，系统需求问题已经提交给产品记录。问题记录台账及运营数量统计及文书统计已经记录完成，现已发送附件给各位领导，请查阅。 '
    # )


user_input = [1, 2, 3, 4, 4, 4]
mailoutput(sheet_ranges, '南沙', user_input)
