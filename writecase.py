import time

from openpyxl import worksheet, workbook, load_workbook
import case


def input_cases():  # 手动输入列表处理
    user_input = [input('新增检查案件\t\t'),
                  input('新增处罚案件\t\t'),
                  input('新增强制案件\t\t'),
                  input('新增检查文书\t\t'),
                  input('新增处罚文书\t\t'),
                  input('新增强制文书\t\t'),
                  input('反馈问题数（个）\t\t'),
                  input(' （Q群）技术支持用户次数\t\t')]

    return user_input


def matching(site, user_input):  # 匹配差值 返回差值
    system_input = case.matching_case(site)
    print('系统案件数')
    print(system_input)
    user_input = user_input[0:6]
    user_input1 = []
    dis = []  # 差值
    if system_input is not None:
        for u in user_input:  # 将列表中字符串转换为数字
            user_input1.append(int(u))
        print('输入案件数')
        print(user_input1)

        try:
            if user_input1 != system_input:
                # print('到这里')
                i = 0
                while i <= 5:
                    d = user_input1[i] - system_input[i]
                    dis.append(d)
                    i += 1
                print('差值')
                print(dis)
                return dis
            else:
                print('数值一样无需添加')
        except IndexError:
            print('今日没有数据')


def add_case(sheet_ranges, site, user_input):  # 将差值添加进统计汇总表
    difference_value = matching(site, user_input)
    max_row = sheet_ranges.max_row
    col = 6
    # print('最大行')
    # print(max_row)
    if difference_value is not None:
        while max_row != 0:

            if sheet_ranges.cell(column=col, row=max_row).value == '有效数':

                if sheet_ranges.cell(column=col + 1, row=2).value == '新增检查案件':
                    sheet_ranges.cell(column=col + 1, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 1, row=max_row).value) + int(difference_value[0])

                if sheet_ranges.cell(column=col + 2, row=2).value == '新增处罚案件':
                    sheet_ranges.cell(column=col + 2, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 2, row=max_row).value) + int(difference_value[1])

                if sheet_ranges.cell(column=col + 3, row=2).value == '新增强制案件':
                    sheet_ranges.cell(column=col + 3, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 3, row=max_row).value) + int(difference_value[2])

                    sheet_ranges.cell(column=col + 4, row=max_row).value = sheet_ranges.cell(column=col + 1,
                                                                                             row=max_row).value \
                                                                           + sheet_ranges.cell(column=col + 2,
                                                                                               row=max_row).value \
                                                                           + sheet_ranges.cell(column=col + 3,
                                                                                               row=max_row).value
                if sheet_ranges.cell(column=col + 5, row=2).value == '新增检查文书':
                    sheet_ranges.cell(column=col + 5, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 5, row=max_row).value) + int(difference_value[3])

                if sheet_ranges.cell(column=col + 6, row=2).value == '新增处罚文书':
                    sheet_ranges.cell(column=col + 6, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 6, row=max_row).value) + int(difference_value[4])

                if sheet_ranges.cell(column=col + 7, row=2).value == '新增强制文书':
                    sheet_ranges.cell(column=col + 7, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 7, row=max_row).value) + int(difference_value[5])

                    sheet_ranges.cell(column=col + 8, row=max_row).value = sheet_ranges.cell(column=col + 5,
                                                                                             row=max_row).value \
                                                                           + sheet_ranges.cell(column=col + 6,
                                                                                               row=max_row).value \
                                                                           + sheet_ranges.cell(column=col + 7,
                                                                                               row=max_row).value

                print('差值注入成功')
                break
            elif sheet_ranges.cell(column=col + 1, row=max_row).value == '有效数':
                col += 1
                if sheet_ranges.cell(column=col + 1, row=2).value == '新增检查案件':
                    sheet_ranges.cell(column=col + 1, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 1, row=max_row).value) + int(difference_value[0])

                if sheet_ranges.cell(column=col + 2, row=2).value == '新增处罚案件':
                    sheet_ranges.cell(column=col + 2, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 2, row=max_row).value) + int(difference_value[1])

                if sheet_ranges.cell(column=col + 3, row=2).value == '新增强制案件':
                    sheet_ranges.cell(column=col + 3, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 3, row=max_row).value) + int(difference_value[2])

                    sheet_ranges.cell(column=col + 4, row=max_row).value = sheet_ranges.cell(column=col + 1,
                                                                                             row=max_row).value \
                                                                           + sheet_ranges.cell(column=col + 2,
                                                                                               row=max_row).value \
                                                                           + sheet_ranges.cell(column=col + 3,
                                                                                               row=max_row).value
                if sheet_ranges.cell(column=col + 5, row=2).value == '新增检查文书':
                    sheet_ranges.cell(column=col + 5, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 5, row=max_row).value) + int(difference_value[3])

                if sheet_ranges.cell(column=col + 6, row=2).value == '新增处罚文书':
                    sheet_ranges.cell(column=col + 6, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 6, row=max_row).value) + int(difference_value[4])

                if sheet_ranges.cell(column=col + 7, row=2).value == '新增强制文书':
                    sheet_ranges.cell(column=col + 7, row=max_row).value = int(
                        sheet_ranges.cell(column=col + 7, row=max_row).value) + int(difference_value[5])

                    sheet_ranges.cell(column=col + 8, row=max_row).value = sheet_ranges.cell(column=col + 5,
                                                                                             row=max_row).value \
                                                                           + sheet_ranges.cell(column=col + 6,
                                                                                               row=max_row).value \
                                                                           + sheet_ranges.cell(column=col + 7,
                                                                                               row=max_row).value

                print('差值注入成功')
                break
            else:
                max_row -= 1
                # print('1')


def write_case(sheet_ranges, place_case, site,
               unit=''):  # place_case 案件总数 user_input 用户输入 site 试点 unit 单位名称
    # filename = '南沙区综合行政执法局执法平台试点台账.xlsx'
    # filename = '惠州市执法平台试点台账.xlsx'

    # ws = wb.active
    """读取指定的sheet表"""
    max_row = sheet_ranges.max_row
    user_view = case.get_page_view(site)  # 登录账号 登录次数
    user_input = input_cases()
    """新增一行，"""
    while True:
        if sheet_ranges['A' + str(max_row)].value == '总计':
            sheet_ranges.insert_rows(max_row)
            break
        max_row -= 1
    # 新增一行后 max_row的值就是当前想操作的这一行

    sheet_ranges['A' + str(max_row)] = max_row - 3  # 序号处理
    sheet_ranges['B' + str(max_row)] = time.strftime("%Y/%m/%d",
                                                     time.localtime())  # 日期处理

    column = 65
    while column < 72:
        if sheet_ranges[chr(column) + str(2)].value == '新增帐号（个）':
            sheet_ranges[chr(column) + str(max_row)] = 0  # 新增账号处理 默认为0
            sheet_ranges[chr(column) + str(max_row + 1)] = str('=SUM('
                                                               + chr(column) + str(4) +  # 起始行
                                                               ':' + chr(column) + str(max_row) + ')')  # 终止行
        if sheet_ranges[chr(column) + str(2)].value == '登录帐号数（个）':
            sheet_ranges[chr(column) + str(max_row)] = user_view[0]  # 访问量处理
            sheet_ranges[chr(column) + str(max_row + 1)] = str('=SUM('
                                                               + chr(column) + str(4) +  # 起始行
                                                               ':' + chr(column) + str(max_row) + ')')  # 终止行

        if sheet_ranges[chr(column) + str(2)].value == '登录次数（次）':
            sheet_ranges[chr(column) + str(max_row)] = user_view[1]  # 访问次数处理
            sheet_ranges[chr(column) + str(max_row + 1)] = str('=SUM('
                                                               + chr(column) + str(4) +  # 起始行
                                                               ':' + chr(column) + str(max_row) + ')')  # 终止行
        column += 1

    """处理惠州单位名称"""
    column = 71  # ASCii码表中71对应G
    if sheet_ranges[chr(column) + str(2)].value != '新增检查案件':  # 防止惠州台账出错
        column += 1
        sheet_ranges['C' + str(max_row)] = unit  # 单位名称处理 只有惠州有
        print('惠州')

    end_column = column + 10
    pc_column = column

    """处理案件总数与删除数"""
    for pc in place_case:
        sheet_ranges[chr(pc_column) + str(max_row + 3)].value = pc
        sheet_ranges[chr(pc_column) + str(max_row + 2)].value = str('=IMSUB('
                                                                    + chr(pc_column) + str(max_row + 1) +
                                                                    '，' + chr(pc_column) + str(max_row + 3) + ')')  # 终
        pc_column += 1

    """处理输入的数据"""
    i = 0
    while column <= end_column:

        asc_column = chr(column)  # ascii码转换字母

        if sheet_ranges[chr(column - 1) + str(2)].value == '新增强制案件' \
                or sheet_ranges[chr(column - 1) + str(2)].value == '新增强制文书':  # 合计的位置 不需要手动输入

            sheet_ranges[asc_column + str(max_row)].value = str('=SUM('
                                                                + chr(column - 3) + str(max_row) +  # 起始行
                                                                ':' + chr(column - 1) + str(max_row) + ')')  # 终止行

            sheet_ranges[asc_column + str(max_row + 1)].value = str('=SUM('
                                                                    + asc_column + str(4) +  # 起始行
                                                                    ':' + asc_column + str(max_row) + ')')  # 终止行

        elif sheet_ranges[chr(column + 1) + str(2)].value == '反馈问题数（个）':  # 办理次数

            sheet_ranges[asc_column + str(max_row)].value = str('=SUM('
                                                                + chr(column - 5) + str(max_row) +  # 起始行
                                                                '+' + chr(column - 1) + str(max_row) + ')')  # 终止行

            sheet_ranges[asc_column + str(max_row + 1)].value = str('=SUM('
                                                                    + asc_column + str(4) +  # 起始行
                                                                    ':' + asc_column + str(max_row) + ')')  # 终止行
        else:

            sheet_ranges[asc_column + str(max_row)].value = int(user_input[i])
            sheet_ranges[asc_column + str(max_row + 1)].value = str('=SUM('
                                                                    + asc_column + str(4) +  # 起始行
                                                                    ':' + asc_column + str(max_row) + ')')
            i += 1

        column += 1
    return user_input
    # add_case(sheet_ranges, site, user_input)
    # wb.save(str('E:/台账/南沙区综合行政执法局执法平台试点台账' + time.strftime("%Y%m%d", time.localtime()) + '.xlsx'))
