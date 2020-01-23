import time

from openpyxl import load_workbook

import Writ
import case
import writeWrit
import writecase
from publicity import write_publicity

wb = load_workbook('南沙区综合行政执法局执法平台试点台账.xlsx')
sheet_ranges = wb['文书统计']  # 文书处理
writeWrit.write_writ(sheet_ranges, 3, 'B', Writ.get_daywrit('A', 'B'))  # 文书处理
writeWrit.write_writ(sheet_ranges, 3, 'C', Writ.get_allwrit('A', 'B'))
# #

sheet_ranges = wb['统计汇总表']  # 统计汇总表处理
user_input = writecase.write_case(sheet_ranges, case.get_case('南沙'), '南沙')
writecase.add_case(sheet_ranges, '南沙', user_input)

wb.save(str('E:/台账/南沙区综合行政执法局执法平台试点台账' +
            time.strftime("%Y%m%d", time.localtime()) + '.xlsx'))
print('南沙台账写入成功')
# #


wb = load_workbook('惠州市执法平台试点台账.xlsx')
sheet_ranges = wb['文书统计']

writeWrit.write_writ(sheet_ranges, 4, 'B', Writ.get_daywrit('C', 'D'))  # 惠州
writeWrit.write_writ(sheet_ranges, 4, 'G', Writ.get_allwrit('C', 'D'))

writeWrit.write_writ(sheet_ranges, 4, 'C', Writ.get_daywrit('E', 'F'))  # 惠城
writeWrit.write_writ(sheet_ranges, 4, 'H', Writ.get_allwrit('E', 'F'))

writeWrit.write_writ(sheet_ranges, 4, 'D', Writ.get_daywrit('G', 'H'))  # 惠阳
writeWrit.write_writ(sheet_ranges, 4, 'I', Writ.get_allwrit('G', 'H'))

writeWrit.write_writ(sheet_ranges, 4, 'E', Writ.get_daywrit('I', 'J'))  # 大亚湾
writeWrit.write_writ(sheet_ranges, 4, 'J', Writ.get_allwrit('I', 'J'))

writeWrit.write_writ(sheet_ranges, 4, 'F', Writ.get_daywrit('K', 'L'))  # 仲恺
writeWrit.write_writ(sheet_ranges, 4, 'K', Writ.get_allwrit('K', 'L'))

sheet_ranges = wb['统计汇总表']  # 统计汇总表处理
print('惠州市城市管理行政执法局大亚湾经济技术开发区分局')
user_input1 = writecase.write_case(sheet_ranges, case.get_case('惠州'), '惠州', '惠州市城市管理行政执法局大亚湾经济技术开发区分局')
print('仲恺高新区城市管理行政执法分局')
user_input2 = writecase.write_case(sheet_ranges, case.get_case('惠州'), '惠州', '仲恺高新区城市管理行政执法分局')
print('惠州市城乡管理和综合执法局')
user_input3 = writecase.write_case(sheet_ranges, case.get_case('惠州'), '惠州', '惠州市城乡管理和综合执法局')
print('惠城区城乡管理和综合执法局')
user_input4 = writecase.write_case(sheet_ranges, case.get_case('惠州'), '惠州', '惠城区城乡管理和综合执法局')
print('惠阳区城乡管理和综合执法局')
user_input5 = writecase.write_case(sheet_ranges, case.get_case('惠州'), '惠州', '惠阳区城乡管理和综合执法局')

i = 0
user_input = []

while i <= 5:
    u = int(user_input1[i]) + int(user_input2[i]) + int(user_input3[i]) + int(user_input4[i]) + int(user_input5[i])
    user_input.append(u)
    i += 1
print(user_input)
writecase.add_case(sheet_ranges, '惠州', user_input)
#
wb.save(str('E:/台账/惠州市执法平台试点台账' +
            time.strftime("%Y%m%d", time.localtime()) + '.xlsx'))  # 测试代码 完成后修改
print('惠州台账写入成功')

write_publicity()  # 执法公示平台

#test