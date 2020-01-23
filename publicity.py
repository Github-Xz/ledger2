import re
import string
import time

from openpyxl import load_workbook

filename = '执法，采集，公示案件文书统计.xlsx'
wb = load_workbook(filename)
ws = wb.active
"""读取指定的sheet表"""
sheet_ranges = wb['02、公示六大类统计']
sheet_ranges1 = wb['03、公示采集统计']


def input_publicity():
    user_input = [input('网上支持单位用户人数：\t\t'),
                  input('客户咨询数量：\t\t'),
                  input('网上咨询数量：\t\t'),
                  input('电话咨询数量:\t\t'),
                  input('已为客户解答数量：\t\t')]
    return user_input


def get_publicity(row, max_row):
    sum1 = []

    while row <= max_row:
        a = sheet_ranges['I' + str(row)].value
        p = re.sub(r'\d{4}-\d{2}-\d{2}', '', a)  # 过滤日期
        number = re.findall(r"\d+", p)  # 提取数字
        for s in number:
            s1 = int(s)
            sum1.append(s1)
        row += 1

    print(sum1)
    return sum1


def get_collect_publicity(row, max_row):
    sum1 = []

    while row <= max_row:
        a = sheet_ranges1['I' + str(row)].value
        # p = re.sub(r'\d{4}-\d{2}-\d{2}', '', a)  # 过滤日期
        number = re.findall(r"\d+", a)  # 提取数字
        for s in number:
            s1 = int(s)
            sum1.append(s1)
        row += 1
    print(sum1)
    return sum1


def mail(row, maxrow):  # 邮件统计
    amd = get_publicity(row, maxrow)
    s = 0
    for a in amd:
        s += a
    return s


def write_publicity():
    filename = '广东省行政执法公示平台数据采集系统运营数量统计.xlsx'
    wb = load_workbook(filename)
    sheet_ranges = wb['运营数量统计']

    col = sheet_ranges.max_column
    t = time.strftime("%m/%d", time.localtime())
    p = re.sub(r'/', '月', t)
    s = str(p + '日')
    year = time.strftime("%Y", time.localtime())
    mon = time.strftime("%m", time.localtime())
    day = time.strftime("%d", time.localtime())
    col += 1
    sheet_ranges.cell(column=col, row=2, value=s)
    sheet_ranges.cell(column=col, row=3, value="当日新增数量")
    sheet_ranges.cell(column=col, row=4, value=0)
    sheet_ranges.cell(column=col, row=5, value=0)

    row = 6
    user_input = input_publicity()
    d = user_input[1]  # 客户咨询数量
    for u in user_input:
        sheet_ranges.cell(column=col, row=row, value=int(u))
        sheet_ranges.cell(column=4, row=row).value += int(u)
        row += 1

    """处理百分比"""
    try:
        subtract = sheet_ranges.cell(column=col, row=row - 1).value - sheet_ranges.cell(
            column=col - 1, row=row - 1).value
        sheet_ranges.cell(column=col, row=row).value = '{:.2%}'.format(subtract / sheet_ranges.cell(
            column=col, row=row - 1).value)
        print(str(sheet_ranges.max_column))
        row += 1
    except ZeroDivisionError:
        sheet_ranges.cell(column=col, row=row).value = 0
        row += 1

    """执法主体和执法部门"""
    amd = get_publicity(34, 35)
    for a in amd:
        sheet_ranges.cell(column=col, row=row, value=int(a))
        row += 1
    row -= 2
    amd = get_collect_publicity(4, 5)
    for a in amd:
        sheet_ranges.cell(column=4, row=row, value=int(a))
        row += 1
    """系统用户数量"""
    row += 1
    amd = get_publicity(37, 37)
    for a in amd:
        b = a - sheet_ranges.cell(column=4, row=row).value
        sheet_ranges.cell(column=col, row=row, value=int(b))
    row += 5
    """公示案件数量："""
    amd = get_publicity(11, 16)
    s = 0
    for a in amd:
        sheet_ranges.cell(column=col, row=row, value=int(a))
        s += a
        row += 1
    sheet_ranges.cell(column=col, row=row - 7, value=int(s))
    row -= 6
    """公示案件数量总数"""
    amd = get_publicity(25, 30)
    s = 0
    for a in amd:
        sheet_ranges.cell(column=4, row=row, value=int(a))
        s += a
        row += 1
    sheet_ranges.cell(column=4, row=row - 7, value=int(s))

    row += 1
    """公示执法清单数量"""
    amd = get_collect_publicity(21, 23)
    s = 0
    for a in amd:
        sheet_ranges.cell(column=col, row=row, value=int(a))
        s += a
        row += 1
    sheet_ranges.cell(column=col, row=row - 4, value=int(s))

    row -= 3
    amd = get_collect_publicity(11, 13)
    s = 0
    for a in amd:
        sheet_ranges.cell(column=4, row=row, value=int(a))
        s += a
        row += 1
    sheet_ranges.cell(column=4, row=row - 4, value=int(s))

    amd = get_collect_publicity(16, 19)
    del amd[1]
    s = 0
    for a in amd:
        sheet_ranges.cell(column=col, row=row, value=int(a))
        s += a
        row += 1
    row -= 3

    amd = get_collect_publicity(6, 9)
    del amd[1]
    s = 0
    for a in amd:
        sheet_ranges.cell(column=4, row=row, value=int(a))
        s += a
        row += 1

    print('各位领导晚上好！\n\n截至' + year + '年' + mon + '月' + day + '日执法信息公示采集系统及公示平台案件录入总量为' + str(mail(18, 23)) + '条记录，' \
                                                                                                            '其中已公示数据总量为' + str(
        mail(25, 30)) + '条。'
                        '行政执法信息公示平台数据采集系统在客户使用过程中问题咨询总量为' + str(sheet_ranges['D7'].value) +
          '条，已为客户解答' + str(d) + '条。\n' +
          '\n本日客户咨询量为' + str(d) + '条，填报案件' + str(mail(4, 9)) + '条，其中已公示案件' + str(mail(11, 16)) +
          '条\n目前系统操作问题都已解决，系统需求问题已经提交给产品记录。问题记录台账及运营数量统计已经记录完成，现已发送附件给各位领导，请查阅。')

    wb.save(str('E:/台账/广东省行政执法公示平台数据采集系统运营数量统计' + time.strftime("%Y%m%d", time.localtime()) + '.xlsx'))
