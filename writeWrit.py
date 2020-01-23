from openpyxl import load_workbook
import Writ


# filename = '惠州市执法平台试点台账.xlsx'
# wb = load_workbook(filename)
# sheet_ranges = wb['文书统计']


def write_writ(sheet_ranges, b_row, column, day_writs):  # b_row 文书名称的开始数, column为文书数量的列

    max_row = sheet_ranges.max_row  # 最大行
    all_day_writs = {}
    row = b_row  # 行数不做限制会不会有啥影响

    """清空文书数量"""
    while b_row <= max_row:

        if sheet_ranges['A' + str(b_row)].value == '合计':
            b_row = row
            print('清空完成')
            break
        sheet_ranges[column + str(b_row)] = 0  # 将数量变成0

        b_row += 1

    """将读取到的文书名称与对应的行数存入字典 all_day_writs"""
    while True:
        str1 = ('A' + str(b_row))  # 文书的名称的列不变 ，变的是每日文书和总文书的列
        if sheet_ranges[str1].value is None:
            break
        else:
            all_day_writs[sheet_ranges[str1].value] = b_row
        b_row += 1

    """将统计分析的文书与南沙文书名称匹配"""
    for day_writ_name, day_writ_num in day_writs.items():  # 注意用items()方法 p88页
        for s, r in all_day_writs.items():
            if day_writ_name == s:
                sheet_ranges[column + str(r)] = day_writ_num


# write_writ(sheet_ranges, 4, 'C', Writ.get_daywrit('E', 'F'))  # 惠城
# write_writ(sheet_ranges, 4, 'H', Writ.get_allwrit('E', 'F'))
#
# wb.save('E:/台账/惠州台账.xlsx')  # 测试代码 完成后修改
# print('惠州台账写入成功')
